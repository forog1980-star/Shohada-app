import csv
import json
import os
import sys
import urllib.request
import urllib.error
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl نصب نیست. اجرا کنید: python -m pip install openpyxl")
    raise SystemExit(1)

SUPABASE_URL = "https://bafrksgdcmglahyrppfy.supabase.co"
SUPABASE_PUBLISHABLE_KEY = "sb_publishable_O5CkSuivysXJf-8hu1IUCA_izu8hWiX"
TABLE = "martyrs"
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "martyrs_master_v2.xlsx")
BATCH_SIZE = 100
EXPECTED_ROWS = 2975


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\u200c", " ").replace("ي", "ی").replace("ى", "ی").replace("ك", "ک")
    return " ".join(text.split())


def digits(value):
    text = clean(value)
    trans = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
    return text.translate(trans)


def value_is_checked(value):
    if value is None:
        return False
    text = clean(value).lower()
    return text not in ("", "0", "false", "no", "خیر", "نه", "-")


def find_col(headers, *names):
    normalized = {clean(h): i for i, h in enumerate(headers)}
    for name in names:
        if clean(name) in normalized:
            return normalized[clean(name)]
    return None


def request_json(method, url, payload=None):
    data = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("apikey", SUPABASE_PUBLISHABLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_PUBLISHABLE_KEY}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "return=minimal")
    try:
        with urllib.request.urlopen(req, timeout=60) as response:
            return response.status, response.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return e.code, body


def main():
    print("=" * 70)
    print("انتقال martyrs_master_v2.xlsx به Supabase")
    print("=" * 70)

    if not os.path.exists(EXCEL_FILE):
        print("ERROR: فایل Excel پیدا نشد:", EXCEL_FILE)
        return 1

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    if wb.sheetnames != ["Sheet1"]:
        print("ERROR: انتظار می‌رود فقط Sheet1 وجود داشته باشد:", wb.sheetnames)
        return 1

    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(x) for x in next(rows)]

    name_col = find_col(headers, "نام")
    lastname_col = find_col(headers, "نام خانوادگی", "نام‌خانوادگی")
    piece_col = find_col(headers, "قطعه")
    row_col = find_col(headers, "ردیف")
    # چون دو ستون با عنوان «ردیف» داریم، ستون محل مزار در فایل جدید ستون پنجم است.
    if headers.count("ردیف") >= 2:
        row_col = [i for i, h in enumerate(headers) if h == "ردیف"][1]
    number_col = find_col(headers, "شماره", "شماره مزار")
    type_col = find_col(headers, "ترمیمی /تعویض", "ترمیمی/تعویض")
    replacement_col = find_col(headers, "تعویضی")
    repair_col = find_col(headers, "مرمتی")
    notes_col = find_col(headers, "ملاحظات")

    stage_cols = [
        ("ترمیمی", "ارسال به واحد مرمت", find_col(headers, "ارسال به واحد مرمت")),
        ("ترمیمی", "سنگ مرمتی آماده", find_col(headers, "سنگ مرمتی آماده")),
        ("ترمیمی", "نصب مرمتی شده", find_col(headers, "نصب سنگ مرمتی آماده شده", "نصب مرمتی شده")),
        ("تعویضی", "ارسال به واحد تعویض", find_col(headers, "ارسال به واحد تعویض")),
        ("تعویضی", "سنگ تعویضی آماده", find_col(headers, "سنگ تعویضی آماده")),
        ("تعویضی", "تعویضی نصب شده", find_col(headers, "نصب سنگ تعویضی آماده شده", "تعویضی نصب شده")),
    ]

    required = {"نام": name_col, "نام خانوادگی": lastname_col, "قطعه": piece_col, "ردیف": row_col, "شماره": number_col}
    missing = [k for k, v in required.items() if v is None]
    if missing:
        print("ERROR: ستون‌های ضروری پیدا نشدند:", ", ".join(missing))
        return 1

    records = []
    duplicate_groups = defaultdict(list)

    for excel_row, row in enumerate(rows, start=2):
        values = list(row)
        if not any(clean(v) for v in values):
            continue

        piece = digits(values[piece_col])
        grave_row = digits(values[row_col])
        grave_number = digits(values[number_col])
        name = clean(values[name_col])
        lastname = clean(values[lastname_col])

        stone_type = ""
        type_value = clean(values[type_col]) if type_col is not None else ""
        if "تعویض" in type_value:
            stone_type = "تعویضی"
        elif "ترمیم" in type_value or "مرمت" in type_value:
            stone_type = "ترمیمی"
        elif replacement_col is not None and value_is_checked(values[replacement_col]):
            stone_type = "تعویضی"
        elif repair_col is not None and value_is_checked(values[repair_col]):
            stone_type = "ترمیمی"

        # اگر چند مرحله علامت خورده باشد، آخرین مرحله ثبت‌شده را به عنوان مرحله فعلی نگه می‌داریم.
        stage = ""
        for candidate_type, candidate_stage, col in stage_cols:
            if col is not None and value_is_checked(values[col]):
                if not stone_type or stone_type == candidate_type:
                    stone_type = candidate_type
                    stage = candidate_stage

        notes = clean(values[notes_col]) if notes_col is not None else ""

        record = {
            "name": name or None,
            "lastname": lastname or None,
            "piece": piece or None,
            "grave_row": grave_row or None,
            "grave_number": grave_number or None,
            "stone_type": stone_type or None,
            "stage": stage or None,
            "notes": notes or None,
            "status": "تأیید شده",
        }
        records.append(record)
        duplicate_groups[(piece, grave_row, grave_number)].append((excel_row, name, lastname))

    wb.close()

    print(f"تعداد رکورد استخراج‌شده: {len(records):,}")
    if len(records) != EXPECTED_ROWS:
        print(f"ERROR: تعداد مورد انتظار {EXPECTED_ROWS:,} نیست. انتقال متوقف شد.")
        return 1

    duplicates = {k: v for k, v in duplicate_groups.items() if k[0] and k[1] and k[2] and len(v) > 1}
    print(f"گروه‌های تکراری قطعه+ردیف+شماره: {len(duplicates):,}")

    report_path = os.path.join(os.path.dirname(__file__), "duplicate_report.csv")
    with open(report_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["قطعه", "ردیف", "شماره", "ردیف Excel", "نام", "نام خانوادگی"])
        for key, items in sorted(duplicates.items()):
            for excel_row, name, lastname in items:
                writer.writerow([*key, excel_row, name, lastname])
    print("گزارش تکراری‌ها:", report_path)

    # نکته: پاک‌سازی بانک از طریق SQL مدیریتی انجام می‌شود؛ این اسکریپت فقط INSERT می‌کند.
    for start in range(0, len(records), BATCH_SIZE):
        batch = records[start:start + BATCH_SIZE]
        status, body = request_json("POST", f"{SUPABASE_URL}/rest/v1/{TABLE}", batch)
        if status < 200 or status >= 300:
            print(f"ERROR در batch {start + 1}-{start + len(batch)}: HTTP {status}")
            print(body)
            return 1
        print(f"ثبت شد: {start + len(batch):,}/{len(records):,}")

    print("=" * 70)
    print("انتقال کامل شد.")
    print("تعداد ثبت‌شده:", f"{len(records):,}")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
