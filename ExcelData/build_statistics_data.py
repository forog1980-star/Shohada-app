import json
import os
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(__file__))
EXCEL_FILE = os.path.join(ROOT, "ExcelData", "martyrs_master_v2.xlsx")
OUT_FILE = os.path.join(ROOT, "frontend", "statistics-data.js")

PIECES = [("17", 103), ("24", 6100), ("26", 4514), ("27", 3177), ("28", 3523), ("29", 2743), ("40", 2948), ("53", 3092)]
STAGES = {
    "ترمیمی": ["ارسال به واحد مرمت", "سنگ مرمتی آماده", "نصب مرمتی شده"],
    "تعویضی": ["ارسال به واحد تعویض", "سنگ تعویضی آماده", "تعویضی نصب شده"],
}

def clean(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().replace("\u200c", " ").replace("ي", "ی").replace("ى", "ی").replace("ك", "ک").split())

def checked(value):
    return clean(value).lower() not in ("", "0", "false", "no", "خیر", "نه", "-")

def norm_piece(value):
    return clean(value).translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")).replace(".0", "")

def find(headers, *names):
    mapping = {clean(h): i for i, h in enumerate(headers)}
    for name in names:
        if clean(name) in mapping:
            return mapping[clean(name)]
    return None

wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
rows = []
for ws in wb.worksheets:
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = [clean(x) for x in next(iterator)]
    except StopIteration:
        continue
    piece_col = find(headers, "قطعه")
    replacement_col = find(headers, "تعویضی")
    repair_col = find(headers, "مرمتی")
    if piece_col is None or (replacement_col is None and repair_col is None):
        continue

    stage_cols = [
        ("ترمیمی", "ارسال به واحد مرمت", find(headers, "ارسال به واحد مرمت")),
        ("ترمیمی", "سنگ مرمتی آماده", find(headers, "سنگ مرمتی آماده")),
        ("ترمیمی", "نصب مرمتی شده", find(headers, "نصب مرمتی شده", "نصب سنگ مرمتی آماده شده")),
        ("تعویضی", "ارسال به واحد تعویض", find(headers, "ارسال به واحد تعویض")),
        ("تعویضی", "سنگ تعویضی آماده", find(headers, "سنگ تعویضی آماده")),
        ("تعویضی", "تعویضی نصب شده", find(headers, "تعویضی نصب شده", "نصب سنگ تعویضی آماده شده")),
    ]

    for values in iterator:
        values = list(values)
        if not any(clean(v) for v in values):
            continue
        piece = norm_piece(values[piece_col])
        if piece not in dict(PIECES):
            continue

        replacement = checked(values[replacement_col]) if replacement_col is not None else False
        repair = checked(values[repair_col]) if repair_col is not None else False
        warning = replacement and repair
        operation_type = "تعویضی" if replacement and not repair else "ترمیمی" if repair and not replacement else ""

        stage = ""
        for stage_type, stage_name, column in stage_cols:
            if column is not None and checked(values[column]):
                if operation_type and stage_type == operation_type:
                    stage = stage_name
                elif not operation_type and not stage:
                    stage = stage_name

        rows.append({"piece": piece, "type": operation_type, "stage": stage, "warning": warning})
wb.close()

model = {
    "pieces": {p: {"total": total, "replacementEligible": 0, "replacementDone": 0, "repairEligible": 0, "repairDone": 0, "warnings": 0} for p, total in PIECES},
    "repair": {"total": 0, "completed": 0, "stages": {s: 0 for s in STAGES["ترمیمی"]}},
    "replacement": {"total": 0, "completed": 0, "stages": {s: 0 for s in STAGES["تعویضی"]}},
    "qualityWarnings": 0,
    "sourceRows": len(rows),
}

for row in rows:
    piece = row["piece"]
    if row["warning"]:
        model["pieces"][piece]["warnings"] += 1
        model["qualityWarnings"] += 1
        continue

    operation_type = row["type"]
    if operation_type not in ("ترمیمی", "تعویضی"):
        continue

    target = model["repair"] if operation_type == "ترمیمی" else model["replacement"]
    target["total"] += 1
    if row["stage"] in target["stages"]:
        target["stages"][row["stage"]] += 1
    if row["stage"] == STAGES[operation_type][2]:
        target["completed"] += 1

    if operation_type == "ترمیمی":
        model["pieces"][piece]["repairEligible"] += 1
        if row["stage"] == STAGES[operation_type][2]:
            model["pieces"][piece]["repairDone"] += 1
    else:
        model["pieces"][piece]["replacementEligible"] += 1
        if row["stage"] == STAGES[operation_type][2]:
            model["pieces"][piece]["replacementDone"] += 1

for piece, item in model["pieces"].items():
    item["replacementRemaining"] = max(item["replacementEligible"] - item["replacementDone"], 0)
    item["repairRemaining"] = max(item["repairEligible"] - item["repairDone"], 0)
    item["eligible"] = item["replacementEligible"] + item["repairEligible"]
    item["done"] = item["replacementDone"] + item["repairDone"]
    item["remaining"] = item["replacementRemaining"] + item["repairRemaining"]
    item["damagePct"] = item["eligible"] / item["total"] * 100 if item["total"] else 0

model["tracked"] = model["repair"]["total"] + model["replacement"]["total"]
model["completed"] = model["repair"]["completed"] + model["replacement"]["completed"]
model["remaining"] = max(model["tracked"] - model["completed"], 0)

with open(OUT_FILE, "w", encoding="utf-8") as handle:
    handle.write("window.GOLZAR_EXCEL_STATS = " + json.dumps(model, ensure_ascii=False, separators=(",", ":")) + ";\n")

print(json.dumps(model, ensure_ascii=False, indent=2))
